#!/usr/bin/env python 
"""
@author:闫学雷
@project:PythonYan
@file: all.py
@time:2021/1/22 0022
"""

import requests
from 自动化测试脚本.接口自动化excel版.uti import ReadExcel_1026
from openpyxl import load_workbook
from 自动化测试脚本.接口自动化excel版.setting import config

import json



class Test_Case():

    def Test_1026(self):
        exceldata = ReadExcel_1026.ReadExcel.getExcelData(1) #调用读取excel的方法
        for i in range(0, len(exceldata)):
            params = exceldata[i] #返回一个字典
            if params['method'] == 'post':
                r = requests.post(url=params['base_url']+params['api'],data=params['requestBody'],headers = eval(params['header'])) #eval方式是把str转dict类型
                # print('接口:',params['base_url']+params['api'])
                try:
                    assert r.status_code == 200
                    wb = load_workbook(config.TEST_CASE_PATH)  # openpyxl只能读取xlsx的问题 而且改后缀的也不能读取
                    sheet = wb['1026']  # 获取sheet
                    response, row = r.status_code,i+2
                    # sheet.cell(行, 列).value = 新的值  # 修改单元格的数据
                    sheet.cell(row, 7).value = response
                    sheet.cell(row, 9).value = 'PASS'
                    sheet.cell(row,10).value = config.report_time
                    wb.save(config.TEST_CASE_PATH)
                    print("写入完成",response,row)

                except:
                    wb = load_workbook(config.TEST_CASE_PATH)  # openpyxl只能读取xlsx的问题 而且改后缀的也不能读取
                    sheet = wb['1026']  # 获取sheet
                    response, row = r.text, i + 2
                    # sheet.cell(行, 列).value = 新的值  # 修改单元格的数据
                    sheet.cell(row, 7).value = response
                    sheet.cell(row, 9).value = 'FAILED'
                    sheet.cell(row,10).value = config.report_time

                    wb.save(config.TEST_CASE_PATH)
                    print("写入完成", response, row)

            else:

                r = requests.get(url=params['base_url']+params['api'],data=params['requestBody'],headers = params['header'])
                try:
                    assert r.status_code == 200
                    wb = load_workbook(config.TEST_CASE_PATH)  # openpyxl只能读取xlsx的问题 而且改后缀的也不能读取
                    sheet = wb['1026']  # 获取sheet
                    response, row = r.status_code, i + 2
                    # sheet.cell(行, 列).value = 新的值  # 修改单元格的数据
                    sheet.cell(row, 7).value = response
                    sheet.cell(row, 9).value = 'PASS'
                    sheet.cell(row,10).value = config.report_time

                    wb.save(config.TEST_CASE_PATH)
                except:
                    wb = load_workbook(config.TEST_CASE_PATH)  # openpyxl只能读取xlsx的问题 而且改后缀的也不能读取
                    sheet = wb['1026']  # 获取sheet
                    response, row = r.text, i + 2
                    # sheet.cell(行, 列).value = 新的值  # 修改单元格的数据
                    sheet.cell(row, 7).value = response
                    sheet.cell(row, 9).value = 'FAILED'
                    sheet.cell(row,10).value = config.report_time

                    wb.save(config.TEST_CASE_PATH)

Test_Case.Test_1026(1)

'''
    ： 一个excel表格文件包含一个工作簿（workbook），一个wb可以包含多个工作表（worksheets）

    用户正在查看的表定义为激活的工作表（active sheet）。每个工作表都有行和列。行以数字1开始，列以字母A开始，

    一个工作表由单元格（cell）组成，cell只存储两种数据类型，数字和字符串（除了纯数字，其它均为字符串类型）

    2： 在excel中设计测试用例的时候，当代码里的值为None的时候，对应cell中不需要输入任何值，空读取出来就是None
    '''

