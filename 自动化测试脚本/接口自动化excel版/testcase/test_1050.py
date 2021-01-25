#!/usr/bin/env python 
"""
@author:闫学雷
@project:PythonYan
@file: test_1050.py
@time:2021/1/25 0025
"""
#!/usr/bin/env python
"""
@author:闫学雷
@project:PythonYan
@file: ReqApi.py
@time:2021/1/22 0022
"""

import requests
from 自动化测试脚本.接口自动化excel版.uti import ReadExcel
from openpyxl import load_workbook
from 自动化测试脚本.接口自动化excel版.setting import config
import time
import json

class Test_Request():

    def test_1050(self):
        exceldata = ReadExcel.ReadExcel.getExcelData1050(1)
        for i in range(0, len(exceldata)):
            params = exceldata[i]
            if params['method'] == 'post':
                # str.encode('utf-8') 转化成utf-8
                r = requests.post(url=params['base_url']+params['api'],data= params['requestBody'].encode('utf-8'),headers = eval(params['header']))
                # print('接口:',params['base_url']+params['api'])
                try:
                    assert r.status_code == 200
                    # return r.status_code
                    Test_Request.writeResult(i+2,r.text,'PASS',config.report_time)
                    print("接口:",params["api_name"],"请求写入完成")
                except:
                    # return r.text
                    Test_Request.writeResult(i+2, r.text, 'FAILED',config.report_time)
                    print("接口:", params["api_name"], "请求写入完成,接口请求失败")
            else:

                r = requests.get(url=params['base_url']+params['api'],data=params['requestBody'],headers = params['header'])
                try:
                    assert r.status_code == 200
                    # return r.status_code
                    print( r.status_code)
                except:
                    print(r.text)

    def writeResult(row,req,result,time):
        wb = load_workbook(config.TEST_CASE_PATH)
        sheet = wb['Sheet1']
        sheet.cell(row, 7).value = req
        sheet.cell(row, 9).value = result
        sheet.cell(row, 10).value = time
        wb.save(config.TEST_CASE_PATH)

# Test_Request.test_case(1)