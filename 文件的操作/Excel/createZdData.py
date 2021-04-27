from openpyxl import load_workbook
import random



# 打开excel
path = "C:\\Users\\86138\\Desktop\\728.xlsx"

workbook1=load_workbook(path)

# 定位单元格
sheet = workbook1['资产导入模板']

# 操作excel的test_data表单 ,定位单元格（cell），根据行列读取测试数据

for i in range(2,12):
    x = random.randint(10000,99999)
    contractName = 'contractName-' + 'xl'+ str(x)
    contractNo = 'contractNo-' + 'xl' + str(x)
    sheet.cell(i,2).value = contractName
    sheet.cell(i,3).value =  contractNo #cell(行，列)

    print(x)

workbook1.save(path)
print('保存成功')
