from openpyxl import load_workbook
from 自动化测试脚本.随机生成四要素 import fourEl
import random



# 打开excel
workbook1=load_workbook("C:\\Users\\Administrator\\Desktop\\sjxs.xlsx")

# 定位单元格
sheet = workbook1['template']

# 操作excel的test_data表单 ,定位单元格（cell），根据行列读取测试数据
a = fourEl('男')

for i in range(2,201):
    x = random.randint(10000,99999)
    # b = a.create_idcard()
    # c = a.create_name()
    # email = b + '@163.com'
    phone = '2570' + str(x)
    z = sheet.cell(i,54).value = phone
    # name = sheet.cell(i,1).value = c  #cell(行，列)

    # print(str(i),'-->',phone)

workbook1.save("C:\\Users\\Administrator\\Desktop\\sjxs.xlsx")
print('保存成功')
