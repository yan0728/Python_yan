#!/usr/bin/env python 
"""
@author:闫学雷
@project:PythonYan
@file: WriteResult.py
@time:2021/1/27 0027
"""

from openpyxl import load_workbook
from 自动化测试脚本.接口自动化excel版.setting import config

def writeResult(api_name,row, req, result, time):
    wb = load_workbook(config.TEST_CASE_PATH)
    sheet = wb[api_name]
    sheet.cell(row, 7).value = req
    sheet.cell(row, 9).value = result
    sheet.cell(row, 10).value = time
    wb.save(config.TEST_CASE_PATH)
